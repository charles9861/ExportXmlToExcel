#!/usr/bin/env python3
"""
export_xml_to_excel.py
Parse clash XML, write an XLSX with columns (1..7) and embed images from XML href.
Usage:
    python export_xml_to_excel.py --xml path/to/report.xml --out Clash_Report.xlsx --max-width 2.0
"""
import argparse
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
from pathlib import Path
from PIL import Image as PILImage
import tempfile
import os
import sys

def find_image_file(href_raw: str, xml_path: Path):
    """Given href from XML, try multiple candidate paths and return Path or None."""
    if not href_raw:
        return None
    href = href_raw.replace("\\", "/").strip()

    candidates = []
    # if it looks absolute (Windows drive letter or starts with /), try it
    p = Path(href)
    if p.is_absolute():
        candidates.append(p)
    # relative to xml file directory
    xml_dir = xml_path.parent
    candidates.append(xml_dir / href)
    # sometimes href is just ELV_files/filename.jpg or only filename; try xml_dir/ELV_files/..., and xml_dir/filename
    if "/" in href:
        candidates.append(xml_dir / Path(href).name)  # just filename in same folder
    else:
        # just a filename, try xml_dir/filename and xml_dir/ELV_files/filename
        candidates.append(xml_dir / href)
        candidates.append(xml_dir / "ELV_files" / href)

    # normalize and check existence
    for cand in candidates:
        try:
            cand_resolved = cand.resolve()
        except Exception:
            cand_resolved = cand
        if cand_resolved.exists():
            return Path(cand_resolved)
    # Not found
    return None

def format_item(clashobject):
    """Read smarttags inside a clashobject and return formatted multiline string."""
    if clashobject is None:
        return ""
    # collect smarttag name->value
    tags = {}
    for st in clashobject.findall(".//smarttag"):
        name_el = st.find("name")
        val_el = st.find("value")
        if name_el is None or val_el is None:
            continue
        name = (name_el.text or "").strip()
        val = (val_el.text or "").strip()
        if name:
            tags[name] = val

    lines = []
    if tags.get("Item Name"):
        lines.append(f"Item Name: {tags['Item Name']}")
    if tags.get("Civil3D General:Network name"):
        lines.append(f"Network: {tags['Civil3D General:Network name']}")
    part = tags.get("Civil3D General:Part Size Name", "")
    itype = tags.get("Item Type", "")
    type_line1 = " ".join(x for x in [part, itype] if x).strip()
    if type_line1:
        lines.append(f"Item Type: {type_line1}")
    inner = tags.get("Civil3D General:Inner Diameter or Width", "")
    outer = tags.get("Civil3D General:Outer Diameter or Width", "")
    if inner or outer:
        lines.append(f"Pipe {inner} x {outer}".strip())
    return "\n".join(lines)

def resize_image_to_temp(orig_path: Path, max_width_px: int):
    """Open image, resize to max_width_px preserving ratio, save to temp file, return temp path."""
    try:
        img = PILImage.open(orig_path)
    except Exception as e:
        print(f"Could not open image {orig_path}: {e}", file=sys.stderr)
        return None
    w, h = img.size
    if w > max_width_px:
        ratio = max_width_px / float(w)
        new_w = int(w * ratio)
        new_h = int(h * ratio)
        img = img.resize((new_w, new_h), PILImage.LANCZOS)
    # save to temp file
    suffix = orig_path.suffix or ".png"
    tf = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    temp_path = Path(tf.name)
    tf.close()
    img.save(temp_path)
    return temp_path

def main(xml_file, out_xlsx, max_width_inch):
    xml_path = Path(xml_file)
    if not xml_path.exists():
        print(f"XML file not found: {xml_file}", file=sys.stderr)
        return

    # convert inches to pixels (approx 96 DPI)
    max_width_px = int(max_width_inch * 96)

    tree = ET.parse(xml_path)
    root = tree.getroot()

    wb = Workbook()
    ws = wb.active
    ws.title = "Clash Report"

    headers = ["RFI No.", "Clash Details", "Clash Entity 1", "Clash Entity 2", "Clash Screenshot", "User Images", "Comments"]
    ws.append(headers)

    temp_files = []  # to delete later

    row_idx = 2
    for i, clash in enumerate(root.findall(".//clashresult"), start=1):
        # Clash details
        clash_name = clash.get("name") or f"Clash {i}"
        distance = clash.get("distance") or "N/A"
        pos = clash.find(".//pos3f")
        if pos is not None:
            x = pos.get("x", "")
            y = pos.get("y", "")
            z = pos.get("z", "")
            coords = f"{x}m,\n{y}m,\n{z}m"
        else:
            coords = "N/A"
        clash_details = f"{clash_name}\nDistance: {distance}m\nClash Point:\n{coords}"

        # Items
        objs = clash.findall(".//clashobject")
        item1 = format_item(objs[0]) if len(objs) > 0 else ""
        item2 = format_item(objs[1]) if len(objs) > 1 else ""

        # write text columns
        ws.cell(row=row_idx, column=1, value=i)
        ws.cell(row=row_idx, column=2, value=clash_details)
        ws.cell(row=row_idx, column=3, value=item1)
        ws.cell(row=row_idx, column=4, value=item2)

        # Column 5: try to find image file and embed
        href_raw = clash.get("href") or ""
        image_path = find_image_file(href_raw, xml_path)
        if image_path:
            # resize and save temp
            temp_img = resize_image_to_temp(image_path, max_width_px)
            if temp_img:
                try:
                    img_for_xlsx = OpenpyxlImage(str(temp_img))
                    # insert at cell E{row_idx}
                    anchor = f"E{row_idx}"
                    ws.add_image(img_for_xlsx, anchor)
                    # adjust row height for image
                    # openpyxl row height uses points; approx: 1 px = 0.75 points
                    img_h_px = PILImage.open(temp_img).size[1]
                    ws.row_dimensions[row_idx].height = int(img_h_px * 0.75) + 6
                    temp_files.append(temp_img)
                except Exception as e:
                    print(f"Failed to embed image {temp_img}: {e}", file=sys.stderr)
                    ws.cell(row=row_idx, column=5, value=str(image_path))
            else:
                ws.cell(row=row_idx, column=5, value=str(image_path))
        else:
            # not found — put the original href text so user can see it
            if href_raw:
                ws.cell(row=row_idx, column=5, value=href_raw)
            else:
                ws.cell(row=row_idx, column=5, value="")

        # Column 6 & 7 placeholders
        ws.cell(row=row_idx, column=6, value="(user images)")
        ws.cell(row=row_idx, column=7, value="")

        row_idx += 1

    # Optionally adjust some column widths
    col_widths = {1:8, 2:40, 3:30, 4:30, 5:22, 6:20, 7:30}
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # Save workbook
    out_path = Path(out_xlsx)
    wb.save(out_path)
    print(f"Saved Excel report: {out_path.resolve()}")

    # clean temporary image files
    for tf in temp_files:
        try:
            os.remove(tf)
        except Exception:
            pass

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Export clash XML to Excel with embedded images.")
    parser.add_argument("--xml", required=True, help="Path to XML file")
    parser.add_argument("--out", default="Clash_Report.xlsx", help="Output XLSX filename")
    parser.add_argument("--max-width", type=float, default=2.0, help="Max image width in inches")
    args = parser.parse_args()
    main(args.xml, args.out, args.max_width)
