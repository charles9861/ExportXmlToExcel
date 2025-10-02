import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
import os
import config  # Import config file with XML_FILE and OUTPUT_FILE


def export_to_excel(xml_file, output_file):
    # Parse XML
    tree = ET.parse(xml_file)
    root = tree.getroot()

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Clash Report"

    # --- Table Header ---
    headers = [
        "ID",            # Column 1
        "Clash Details", # Column 2
        "Item 1",        # Column 3
        "Item 2",        # Column 4
        "Clash Image",   # Column 5
        "User Images",   # Column 6
        "Comments"       # Column 7
    ]
    ws.append(headers)

    clash_id = 1

    # Loop through clashgroups
    groups = root.findall(".//clashgroup")
    print(f"Found {len(groups)} clash groups")

    for clash_group in groups:
        group_name = clash_group.get("name", "Unknown Group")
        results = clash_group.findall(".//clashresult")
        print(f"  Group '{group_name}' has {len(results)} results")

        for clash in results:
            row = clash_id + 1

            # -----------------------
            # Column 1: ID
            # -----------------------
            ws.cell(row=row, column=1, value=clash_id)

            # -----------------------
            # Column 2: Clash Details
            # -----------------------
            clash_name = clash.get("name") or f"Clash {clash_id}"
            distance = clash.get("distance") or "N/A"

            pos = clash.find("pos3f")
            coords = ""
            if pos is not None:
                x = f"{float(pos.get('x')):.3f}m,"
                y = f"{float(pos.get('y')):.3f}m,"
                z = f"{float(pos.get('z')):.3f}m"
                coords = f"\n{x}\n{y}\n{z}"

            objs = clash.findall("clashobject")
            item1_name = get_item_name(objs[0]) if len(objs) > 0 else "Unknown"
            item2_name = get_item_name(objs[1]) if len(objs) > 1 else "Unknown"

            clash_details = (
                f"Clash Group: {group_name}\n"
                f"{clash_name}\n"
                f"Between {item1_name} and {item2_name}\n"
                f"Distance: {distance}m\n"
                f"Clash Point:{coords}"
            )
            ws.cell(row=row, column=2, value=clash_details)

            # -----------------------
            # Column 3 & 4: Item Details
            # -----------------------
            item1_details = get_item_details(objs[0]) if len(objs) > 0 else "Unknown"
            item2_details = get_item_details(objs[1]) if len(objs) > 1 else "Unknown"

            ws.cell(row=row, column=3, value=item1_details)
            ws.cell(row=row, column=4, value=item2_details)

            # -----------------------
            # Column 5: Clash Image
            # -----------------------
            img_href_raw = clash.get("href") or ""
            img_href = img_href_raw.replace("\\", "/").strip()
            if img_href and os.path.exists(img_href):
                img = XLImage(img_href)
                img.width, img.height = 200, 150
                ws.add_image(img, f"E{row}")
            else:
                ws.cell(row=row, column=5, value="No image")

            ws.cell(row=row, column=6, value="Add images manually")
            ws.cell(row=row, column=7, value="")

            clash_id += 1

    # Save Excel file
    wb.save(output_file)
    print(f"✅ Export complete: {output_file}")


def get_item_name(clashObject):
    if clashObject is None:
        return "Unknown"
    for tag in clashObject.findall("smarttag"):
        if tag.findtext("name", "").strip() == "Item Name":
            return tag.findtext("value", "Unknown").strip()
    return "Unknown"


def get_item_details(clashObject):
    if clashObject is None:
        return "Unknown"
    smarttags = clashObject.findall("smarttag")
    data = {}
    for tag in smarttags:
        name = tag.findtext("name", "").strip()
        value = tag.findtext("value", "").strip()
        if not name or not value:
            continue
        if name == "Item Name":
            data["itemName"] = value
        elif name == "Civil3D General:Network name":
            data["network"] = value
        elif name == "Civil3D General:Part Size Name":
            data["partSize"] = value
        elif name == "Item Type":
            data["itemType"] = value
        elif name == "Civil3D General:Inner Diameter or Width":
            data["inner"] = value
        elif name == "Civil3D General:Outer Diameter or Width":
            data["outer"] = value
    typeLine1 = " ".join(filter(None, [data.get("partSize"), data.get("itemType")]))
    typeLine2 = ""
    if data.get("inner") or data.get("outer"):
        typeLine2 = f"Pipe {data.get('inner','')} x {data.get('outer','')}".strip()
    lines = []
    if data.get("itemName"): lines.append(f"Item Name: {data['itemName']}")
    if data.get("network"): lines.append(f"Network: {data['network']}")
    if typeLine1: lines.append(f"Item Type: {typeLine1}")
    if typeLine2: lines.append(typeLine2)
    return "\n".join(lines)


if __name__ == "__main__":
    export_to_excel(config.XML_FILE, config.OUTPUT_FILE)
