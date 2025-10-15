#!/usr/bin/env python3
"""
python export_xml_to_excel_v6.py
Adds "Item 2 Name" column to Clash Report.
"""

import config
import xml.etree.ElementTree as ET
from pathlib import Path
import os
import tempfile
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from PIL import Image as PILImage

# ---------- Layout constants ----------
HEADER_HEIGHT = 35
DATA_ROW_HEIGHT = 200
HEADER_FILL = "2C7676"
HEADER_FONT_COLOR = "FFFFFF"
ALT_ROW_FILL = "DCE6F1"
TABLE_NAME = "Clash_1"

# Updated column widths (8 columns now)
COL_WIDTHS = {
    1: 5,   # A RFI No
    2: 30,  # B Clash Details
    3: 30,  # C Item1
    4: 30,  # D Item2
    5: 25,  # E Item2 Name
    6: 45,  # F Clash Image
    7: 45,  # G User Images
    8: 30   # H Comments
}

IMAGE_PADDING_PX = 8

# ---------- Helper functions ----------
def col_width_to_pixels(col_width):
    return int(col_width * 7 + 5)

def row_height_to_pixels(row_height_pts):
    return int(row_height_pts * 96 / 72)

def find_image_file(href_raw: str, xml_path: Path):
    if not href_raw:
        return None
    href = href_raw.replace("\\", "/").strip()
    candidates = []
    p = Path(href)
    if p.is_absolute():
        candidates.append(p)
    xml_dir = xml_path.parent
    candidates.append(xml_dir / href)
    fname = Path(href).name
    candidates.append(xml_dir / fname)
    candidates.append(xml_dir / "ELV_files" / fname)
    for cand in candidates:
        try:
            cand_resolved = cand.resolve()
        except Exception:
            cand_resolved = cand
        if cand_resolved.exists():
            return Path(cand_resolved)
    return None

def resize_image_for_cell(orig_path: Path, target_width_px: int, target_height_px: int, padding_px=8):
    try:
        img = PILImage.open(orig_path)
    except Exception as e:
        print(f"Could not open image {orig_path}: {e}")
        return None
    max_w = max(1, target_width_px - padding_px)
    max_h = max(1, target_height_px - padding_px)
    w, h = img.size
    ratio = min(max_w / w, max_h / h, 1.0)
    new_w = int(w * ratio)
    new_h = int(h * ratio)
    img = img.resize((new_w, new_h), PILImage.LANCZOS)
    suffix = orig_path.suffix or ".png"
    tf = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    tmp_path = Path(tf.name)
    tf.close()
    img.save(tmp_path)
    return tmp_path

def get_item_name_short(item_details_text):
    for line in item_details_text.splitlines():
        if line.startswith("Item Name:"):
            return line.replace("Item Name:", "").strip()
    first = item_details_text.splitlines()[0] if item_details_text else ""
    return first or "Unknown"

def get_item_details(clash_object):
    if clash_object is None:
        return ""
    tags = {}
    for st in clash_object.findall(".//smarttag"):
        name = st.findtext("name", "").strip()
        val = st.findtext("value", "").strip()
        if name:
            tags[name] = val
    lines = []
    if tags.get("Item Name"):
        lines.append(f"Item Name: {tags.get('Item Name')}")
    if tags.get("Civil3D General:Network name"):
        lines.append(f"Network: {tags.get('Civil3D General:Network name')}")
    part = tags.get("Civil3D General:Part Size Name", "")
    itype = tags.get("Item Type", "")
    type_line = " ".join(filter(None, [part, itype])).strip()
    if type_line:
        lines.append(f"Item Type: {type_line}")
    inner = tags.get("Civil3D General:Inner Diameter or Width", "")
    outer = tags.get("Civil3D General:Outer Diameter or Width", "")
    if inner or outer:
        lines.append(f"Pipe {inner} x {outer}".strip())
    return "\n".join(lines)

# ---------- Main export function ----------
def export_to_excel(xml_file, output_file):
    xml_path = Path(xml_file)
    if not xml_path.exists():
        print(f"XML file not found: {xml_file}")
        return

    tree = ET.parse(xml_path)
    root = tree.getroot()

    wb = Workbook()
    ws = wb.active
    ws.title = "Clash Report"

    for col_idx, w in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # Updated header
    headers = ["RFI No", "Clash Details", "Item 1", "Item 2", "Item 2 Name",
               "Clash Image", "User Images", "Comments"]
    ws.append(headers)

    ws.row_dimensions[1].height = HEADER_HEIGHT
    header_fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
    header_font = Font(color=HEADER_FONT_COLOR, bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    start_data_row = 2
    current_row = start_data_row
    temp_images = []

    clashes = root.findall(".//clashresult")
    for i, clash in enumerate(clashes, start=1):
        clash_group = "Unknown Group"
        clash_guid = clash.get("guid")
        for test in root.findall(".//clashtest"):
            for cr in test.findall(".//clashresult"):
                if clash_guid and cr.get("guid") == clash_guid:
                    clash_group = test.get("name", "Unknown Group")
                    break
            if clash_group != "Unknown Group":
                break

        clash_name = clash.get("name", f"Clash{i}")
        distance = clash.get("distance", "N/A")
        pos = clash.find(".//pos3f")
        coords_text = ""
        x_val = y_val = z_val = None
        if pos is not None:
            try:
                x_val = float(pos.get("x") or 0.0)
                y_val = float(pos.get("y") or 0.0)
                z_val = float(pos.get("z") or 0.0)
                coords_text = f"{x_val:.3f}m,\n{y_val:.3f}m,\n{z_val:.3f}m"
            except Exception:
                coords_text = ""

        objs = clash.findall(".//clashobject")
        item1_text = get_item_details(objs[0]) if len(objs) > 0 else ""
        item2_text = get_item_details(objs[1]) if len(objs) > 1 else ""
        item2_name = get_item_name_short(item2_text)

        between_line = f"Between: {get_item_name_short(item1_text)} and {item2_name}\n"
        clash_details = (
            f"Clash Group: {clash_group}\n"
            f"{between_line}\n"
            f"{clash_name}\n"
            f"Distance: {distance}m\n"
            f"Clash Point:\n{coords_text}"
        )

        # ---- write to Excel ----
        ws.cell(row=current_row, column=1, value=i)
        ws.cell(row=current_row, column=2, value=clash_details)
        ws.cell(row=current_row, column=3, value=item1_text)
        ws.cell(row=current_row, column=4, value=item2_text)
        ws.cell(row=current_row, column=5, value=item2_name)

        href_raw = clash.get("href") or ""
        img_path = find_image_file(href_raw, xml_path)
        if img_path:
            col_w = COL_WIDTHS[6]
            target_w_px = col_width_to_pixels(col_w)
            target_h_px = row_height_to_pixels(DATA_ROW_HEIGHT)
            tmp_img = resize_image_for_cell(img_path, target_w_px, target_h_px, padding_px=IMAGE_PADDING_PX)
            if tmp_img:
                try:
                    img_obj = XLImage(str(tmp_img))
                    anchor_cell = f"{get_column_letter(6)}{current_row}"
                    ws.add_image(img_obj, anchor_cell)
                    temp_images.append(tmp_img)
                except Exception as err:
                    ws.cell(row=current_row, column=6, value=str(img_path))
            else:
                ws.cell(row=current_row, column=6, value=str(img_path))
        else:
            ws.cell(row=current_row, column=6, value=href_raw or "")

        ws.cell(row=current_row, column=7, value="(user images)")
        ws.cell(row=current_row, column=8, value="")

        ws.row_dimensions[current_row].height = DATA_ROW_HEIGHT

        offset = current_row - start_data_row
        for c in range(1, 9):
            cell = ws.cell(row=current_row, column=c)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.font = Font(color="000000")
            if (offset % 2) == 0:
                cell.fill = PatternFill(start_color=ALT_ROW_FILL, end_color=ALT_ROW_FILL, fill_type="solid")

        current_row += 1

    # ---------- Table ----------
    last_row = current_row - 1
    if last_row >= start_data_row:
        table_ref = f"A1:H{last_row}"
        table = Table(displayName=TABLE_NAME, ref=table_ref)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=False, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)

    # ---------- Borders ----------
    thick = Side(style="thick", color="000000")
    double = Side(style="double", color="000000")
    for r in range(1, last_row + 1):
        for c in range(1, 9):
            if r == 1:
                top = thick
            else:
                top = double
            bottom = thick if r == last_row else double
            left = thick if c == 1 else double
            right = thick if c == 8 else double
            ws.cell(row=r, column=c).border = Border(left=left, right=right, top=top, bottom=bottom)

    # ---------- Clash Points Sheet ----------
    cp = wb.create_sheet(title="Clash_Points")
    cp_headers = ["ID", "Group", "Clash Name", "X", "Y", "Z"]
    cp.append(cp_headers)
    cp.row_dimensions[1].height = HEADER_HEIGHT
    for ci in range(1, len(cp_headers) + 1):
        cell = cp.cell(row=1, column=ci)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    row_index = 2
    for i, clash in enumerate(clashes, start=1):
        clash_group = "Unknown Group"
        for test in root.findall(".//clashtest"):
            if clash in test.findall(".//clashresult"):
                clash_group = test.get("name", "Unknown Group")
                break
        clash_name = clash.get("name", f"Clash{i}")
        pos = clash.find(".//pos3f")
        if pos is not None:
            x_val = float(pos.get("x") or 0.0)
            y_val = float(pos.get("y") or 0.0)
            z_val = float(pos.get("z") or 0.0)
        else:
            x_val = y_val = z_val = None
        cp.cell(row=row_index, column=1, value=i)
        cp.cell(row=row_index, column=2, value=clash_group)
        cp.cell(row=row_index, column=3, value=clash_name)
        if x_val is not None:
            cp.cell(row=row_index, column=4, value=round(x_val, 3))
            cp.cell(row=row_index, column=5, value=round(y_val, 3))
            cp.cell(row=row_index, column=6, value=round(z_val, 3))
        row_index += 1

    cp.column_dimensions['A'].width = 6
    cp.column_dimensions['B'].width = 18
    cp.column_dimensions['C'].width = 25
    cp.column_dimensions['D'].width = 12
    cp.column_dimensions['E'].width = 12
    cp.column_dimensions['F'].width = 12

    wb.save(output_file)
    print(f"Saved: {output_file}")

    for t in temp_images:
        try:
            os.remove(t)
        except Exception:
            pass

if __name__ == "__main__":
    export_to_excel(config.XML_FILE, config.OUTPUT_FILE)
